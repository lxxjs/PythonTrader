#2022-11-26

"""
- 기업가치 = 자기자본 + (초과이익/할인율)
: 자기자본 = 지배기업소유주지분(지배주주지분)
: 할인율 = BBB- 등급 회사채 5년 수익률 (매일 갱신) https://www.kisrating.com/ratingsStatistics/statics_spread.do
: 초과이익 = 자기자본 * (가중평균 3년 ROE - 할인율)
- 적정주가 = 기업가치 / 유통주식수

"""
from urllib import parse

import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd

# 전역 변수 선언
## 할인율
discountRateLocation: int  = 87 #BBB- 5y
KISlink: str = 'https://www.kisrating.com/ratingsStatistics/statics_spread.do'

## 다트 재무제표
DART_API_Key = 'b64695f3f2a79d07bde772ffa630f935e0d050c0'



def getDiscountRate(link, location) -> float:
    html = requests.get(link).text
    soup = BeautifulSoup(html, "html.parser")
    result = float(soup.select('td.ar.pr40')[location].text)
    return result

a = getDiscountRate(KISlink, discountRateLocation)
print("Today's discount rate :", a)


'''
#res = requests.get('https://opendart.fss.or.kr/api/company.json', params={'crtfc_key' : api_key, 'corp_code' : samsung})
res = requests.get('https://www.fnspace.com/Api/FinanceApi?key=sample&format=json&code=A005930&item=M111000,M111100,M111600,M113000&consolgb=M&annualgb=A&fraccyear=2014&toaccyear=2018')
res.raise_for_status()
print(res)
data = res.json()
print(data)
'''

# 주식 DB 읽기
workbook = openpyxl.load_workbook("/workspace/PythonTrader/stocksDB_edit.xlsx")
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column

'''
## 주식 명으로 주식코드 가져오기
target_name = input("주식명: ").upper()


# for row in range(2, rows + 1):
#     comp_name = sheet.cell(row, 2).value.upper()
#     if comp_name == target_name:
#         print(target_name, ":", sheet.cell(row, 1).value)

'''




def get_fnguide(code):
    get_param = {
        'pGB':1,
        'gicode':'A%s'%(code),
        'cID':'',
        'MenuYn':'Y',
        'ReportGB':'',
        'NewMenuID':101,
        'stkGb':701,
    }
    get_param = parse.urlencode(get_param)
    url="http://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?%s"%(get_param)
    tables = pd.read_html(url, header=0)
    return(tables)


def get_equity_capital(code):
    if len(get_fnguide(code)) > 10 and get_fnguide(code)[11].iloc[10][5]:
        annual = get_fnguide(code)[11] # 연결-연간 재무제표
        return annual.iloc[10][5]
    return 0

def get_roe(code):
    annual = get_fnguide(code)[11] # 연결-연간 재무제표
    return annual.iloc[18][3:6].tolist() # 최근 3개년 확정 ROE

def get_roe_weighted_mean(code):
    roes = get_roe(code)
    sum = 0
    for i in range(2,len(roes)):
        sum += float(roes[i])
    return sum / 3


for i in range(826,rows + 1):
    this_cell = sheet.cell(i,4)
    code = sheet.cell(i,1).value
    this_cell.value = get_equity_capital(code)
    print(sheet.cell(i,2).value, ":", this_cell.value)
    workbook.save("/workspace/PythonTrader/stocksDB_edit.xlsx")

workbook.close()

print(get_roe_weighted_mean("005930"))
print(get_equity_capital("005930"))
